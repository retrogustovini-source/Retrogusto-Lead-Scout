import subprocess
import sys

subprocess.check_call([sys.executable, "-m", "pip", "install", "httpx==0.27.0", "openpyxl==3.1.2", "psycopg2-binary==2.9.9", "-q"])

import os
import re
import csv
import asyncio
import logging
from datetime import date
import httpx
import psycopg2
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
log = logging.getLogger(__name__)

GOOGLE_API_KEY = os.environ["GOOGLE_API_KEY"]
TELEGRAM_TOKEN = os.environ["TELEGRAM_TOKEN"]
TELEGRAM_CHAT_ID = os.environ["TELEGRAM_CHAT_ID"]
XLSX_PATH = os.environ.get("XLSX_PATH", "leads.xlsx")
DATABASE_URL = os.environ["DATABASE_URL"]

SEARCH_QUERIES = [
    ("wine bar Praha 1", 50.0878, 14.4205, 3000),
    ("wine bar Praha 2", 50.0750, 14.4378, 3000),
    ("wine bar Praha 3", 50.0833, 14.4667, 3000),
    ("wine bar Praha 4", 50.0400, 14.4500, 4000),
    ("wine bar Praha 5", 50.0667, 14.3833, 4000),
    ("wine bar Praha 6", 50.1000, 14.3833, 4000),
    ("wine bar Praha 7", 50.1000, 14.4333, 3000),
    ("wine bar Praha 8", 50.1167, 14.4667, 4000),
    ("wine bar Praha 9", 50.1333, 14.5000, 4000),
    ("wine bar Praha 10", 50.0667, 14.5000, 4000),
    ("vinoteka Praha 1", 50.0878, 14.4205, 3000),
    ("vinoteka Praha 2", 50.0750, 14.4378, 3000),
    ("vinoteka Praha 3", 50.0833, 14.4667, 3000),
    ("vinoteka Praha 4", 50.0400, 14.4500, 4000),
    ("vinoteka Praha 5", 50.0667, 14.3833, 4000),
    ("enoteca Praha", 50.0755, 14.4378, 15000),
    ("vinarna Praha", 50.0755, 14.4378, 15000),
    ("ristorante italiano Praha 1", 50.0878, 14.4205, 3000),
    ("ristorante italiano Praha 2", 50.0750, 14.4378, 3000),
    ("ristorante italiano Praha 3", 50.0833, 14.4667, 3000),
    ("restaurant wine list Praha", 50.0755, 14.4378, 15000),
    ("bistro wine Praha", 50.0755, 14.4378, 15000),
    ("degustace vina Praha", 50.0755, 14.4378, 15000),
    ("wine bar Kladno", 50.1479, 14.1011, 8000),
    ("vinoteka Kladno", 50.1479, 14.1011, 8000),
    ("wine bar Beroun", 49.9603, 14.0706, 6000),
    ("vinoteka Beroun", 49.9603, 14.0706, 6000),
    ("wine bar Melnik", 50.3500, 14.4833, 6000),
    ("vinoteka Melnik", 50.3500, 14.4833, 6000),
    ("wine bar Mlada Boleslav", 50.4131, 14.9058, 6000),
    ("wine bar Kolin", 50.0269, 15.2000, 6000),
    ("vinoteka Kolin", 50.0269, 15.2000, 6000),
    ("wine bar Ricany", 49.9931, 14.6556, 5000),
    ("wine bar Benesov", 49.7817, 14.6858, 5000),
    ("vinoteka Benesov", 49.7817, 14.6858, 5000),
    ("wine bar Brandys nad Labem", 50.1833, 14.6667, 5000),
    ("wine bar Neratovice", 50.2592, 14.5178, 5000),
    ("ristorante italiano Kladno", 50.1479, 14.1011, 8000),
    ("ristorante italiano Beroun", 49.9603, 14.0706, 6000),
    ("ristorante italiano Melnik", 50.3500, 14.4833, 6000),
]

WINE_SIGNALS_NAME = [
    "wine", "vino", "vinoteka", "vinarna", "enoteca", "wein",
    "vinice", "sommelier", "cava", "cellar", "sklep", "degustace",
    "bacchus", "vinný", "vinné", "enolog", "vinařství", "prosecco",
    "champagne", "barrique", "cantina", "osteria", "trattoria",
    "ristorante", "italiano", "italiana", "bistro", "brasserie",
    "vinotéka", "vinárna",
]

WINE_SIGNALS_TYPES = {"bar", "night_club"}

EXCLUDE_KEYWORDS = [
    "supermarket", "tesco", "albert", "billa", "lidl", "penny", "kaufland", "spar",
    "mcdonald", "kfc", "subway", "burger king", "pizza hut", "domino",
    "kebab", "kebap", "gyros", "karaoke",
    "cukrarna", "cukrárna", "pekarna", "pekárna", "pastry", "bakery", "cake", "cakes", "dorty",
    "pivnice", "pivovar", "brewery", "birreria", "beerhouse", "beer house",
    "hospoda", "hospůdka", "pub ",
    "benzinka", "čerpací", "gas station",
    "lékárna", "pharmacy", "fitness", "gym",
]


def normalize_name(name: str) -> str:
    return re.sub(r'\s+', ' ', name.lower().strip())


def has_wine_signal(name: str, types: list) -> bool:
    name_lower = name.lower()
    if any(kw in name_lower for kw in WINE_SIGNALS_NAME):
        return True
    return any(t in WINE_SIGNALS_TYPES for t in types)


def has_exclude_keyword(name: str) -> bool:
    name_lower = name.lower()
    return any(kw in name_lower for kw in EXCLUDE_KEYWORDS)


def get_db_conn():
    return psycopg2.connect(DATABASE_URL)


def init_db():
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS found_leads (
            id SERIAL PRIMARY KEY,
            name_normalized TEXT UNIQUE NOT NULL,
            name_original TEXT NOT NULL,
            found_date DATE NOT NULL DEFAULT CURRENT_DATE
        )
    """)
    conn.commit()
    cur.close()
    conn.close()
    log.info("DB inizializzato")


def load_known_leads_from_db() -> set:
    conn = get_db_conn()
    cur = conn.cursor()
    cur.execute("SELECT name_normalized FROM found_leads")
    rows = cur.fetchall()
    cur.close()
    conn.close()
    return {row[0] for row in rows}


def save_leads_to_db(new_leads: list):
    conn = get_db_conn()
    cur = conn.cursor()
    for lead in new_leads:
        try:
            cur.execute(
                "INSERT INTO found_leads (name_normalized, name_original) VALUES (%s, %s) ON CONFLICT DO NOTHING",
                (normalize_name(lead["name"]), lead["name"])
            )
        except Exception as e:
            log.warning(f"DB insert error for {lead['name']}: {e}")
    conn.commit()
    cur.close()
    conn.close()
    log.info(f"Salvati {len(new_leads)} lead nel DB")


def load_existing_leads_from_xlsx(xlsx_path: str) -> set:
    existing = set()
    if not os.path.exists(xlsx_path):
        log.warning(f"XLSX not found at {xlsx_path}")
        return existing
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["Leads"]
    first_row = True
    for row in ws.iter_rows(values_only=True):
        if first_row:
            first_row = False
            continue
        if row[0]:
            existing.add(normalize_name(str(row[0])))
    wb.close()
    return existing


async def search_places(client: httpx.AsyncClient, query: str, lat: float, lng: float, radius: int) -> list:
    url = "https://places.googleapis.com/v1/places:searchText"
    headers = {
        "Content-Type": "application/json",
        "X-Goog-Api-Key": GOOGLE_API_KEY,
        "X-Goog-FieldMask": (
            "places.displayName,places.formattedAddress,"
            "places.nationalPhoneNumber,places.websiteUri,"
            "places.types,places.rating,places.userRatingCount,places.id"
        )
    }
    payload = {
        "textQuery": query,
        "languageCode": "cs",
        "maxResultCount": 20,
        "locationBias": {
            "circle": {
                "center": {"latitude": lat, "longitude": lng},
                "radius": float(radius)
            }
        }
    }
    try:
        resp = await client.post(url, headers=headers, json=payload, timeout=15)
        resp.raise_for_status()
        return resp.json().get("places", [])
    except Exception as e:
        log.error(f"Error searching '{query}': {e}")
        return []


def extract_zone(address: str, query: str) -> str:
    if not address:
        return "praga"
    address_lower = address.lower()
    cities = {
        "kladno": "Kladno", "beroun": "Beroun", "mělník": "Mělník", "melnik": "Mělník",
        "mladá boleslav": "Mladá Boleslav", "kolín": "Kolín", "kolin": "Kolín",
        "říčany": "Říčany", "ricany": "Říčany", "benešov": "Benešov", "benesov": "Benešov",
        "brandýs": "Brandýs n.L.", "neratovice": "Neratovice",
    }
    for key, val in cities.items():
        if key in address_lower or key in query.lower():
            return val
    match = re.search(r'praha\s*(\d+)', address_lower)
    if match:
        return f"praga {match.group(1)}"
    districts = [
        "vinohrady", "žižkov", "smíchov", "dejvice", "holešovice",
        "nusle", "vršovice", "karlín", "josefov", "letná", "bubeneč", "anděl", "pankrác"
    ]
    for d in districts:
        if d in address_lower:
            return d
    return "praga"


def guess_type(types: list, name: str) -> str:
    name_lower = name.lower()
    if any(w in name_lower for w in ["vinoteka", "vinarna", "enoteca", "wine bar", "vinný bar", "vinotéka", "vinárna"]):
        return "Wine Bar"
    if any(w in name_lower for w in ["trattoria", "ristorante", "osteria", "italiano", "italiana"]):
        return "Ristorante IT"
    if "restaurant" in types:
        return "Ristorante"
    if "bar" in types:
        return "Wine Bar"
    return "Ristorante"


def save_leads_to_csv(new_leads: list) -> str:
    today_str = date.today().strftime("%Y-%m-%d")
    output_path = f"nuovi_lead_{today_str}.csv"
    today_display = date.today().strftime("%d/%m/%Y")
    headers = [
        "Lead Name", "Type", "Language", "Zone", "Contact Channels",
        "Contact Quality", "Situation", "Status", "Priority",
        "Date Added", "Next Action", "Probability (%)",
        "email", "telefono", "website", "rating", "recensioni"
    ]
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for lead in new_leads:
            writer.writerow([
                lead["name"], lead["type"], "CZ", lead["zone"],
                "Email", "freddo", "cold mail", "Waiting", "Media",
                today_display, "Email follow-up", 0.2,
                "", lead.get("phone", ""), lead.get("website", ""),
                lead.get("rating", ""), lead.get("reviews", ""),
            ])
    log.info(f"CSV salvato: {output_path}")
    return output_path


async def send_telegram_message(client: httpx.AsyncClient, text: str):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    resp = await client.post(url, json={"chat_id": TELEGRAM_CHAT_ID, "text": text, "parse_mode": "HTML"}, timeout=15)
    resp.raise_for_status()


async def send_telegram_file(client: httpx.AsyncClient, file_path: str, caption: str):
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendDocument"
    with open(file_path, "rb") as f:
        files = {"document": (os.path.basename(file_path), f, "text/csv")}
        data = {"chat_id": TELEGRAM_CHAT_ID, "caption": caption}
        resp = await client.post(url, data=data, files=files, timeout=30)
        resp.raise_for_status()


async def main():
    log.info("🍷 Retrogusto Lead Scout avviato")
    init_db()

    xlsx_leads = load_existing_leads_from_xlsx(XLSX_PATH)
    db_leads = load_known_leads_from_db()
    known_leads = xlsx_leads | db_leads
    log.info(f"Lead noti: {len(xlsx_leads)} xlsx + {len(db_leads)} DB = {len(known_leads)} totali")

    new_leads = []
    seen_place_ids = set()

    async with httpx.AsyncClient() as client:
        for (query, lat, lng, radius) in SEARCH_QUERIES:
            log.info(f"Ricerca: {query}")
            places = await search_places(client, query, lat, lng, radius)
            log.info(f"  → {len(places)} risultati")
            for place in places:
                place_id = place.get("id", "")
                if place_id in seen_place_ids:
                    continue
                seen_place_ids.add(place_id)
                name = place.get("displayName", {}).get("text", "")
                if not name:
                    continue
                if has_exclude_keyword(name):
                    log.info(f"  → escluso: {name}")
                    continue
                if normalize_name(name) in known_leads:
                    log.info(f"  → già noto: {name}")
                    continue
                types = place.get("types", [])
                if not has_wine_signal(name, types):
                    log.info(f"  → no wine signal: {name}")
                    continue
                lead = {
                    "name": name,
                    "type": guess_type(types, name),
                    "zone": extract_zone(place.get("formattedAddress", ""), query),
                    "phone": place.get("nationalPhoneNumber", ""),
                    "website": place.get("websiteUri", ""),
                    "rating": place.get("rating", ""),
                    "reviews": place.get("userRatingCount", ""),
                }
                new_leads.append(lead)
                log.info(f"  ✅ {name} ({lead['type']}, {lead['zone']})")
            await asyncio.sleep(1)

    log.info(f"Nuovi lead trovati: {len(new_leads)}")

    async with httpx.AsyncClient() as client:
        if not new_leads:
            await send_telegram_message(client, "🍷 <b>Retrogusto Lead Scout</b>\n\nNessun nuovo lead trovato questa settimana.")
            return

        save_leads_to_db(new_leads)
        today_str = date.today().strftime("%d/%m/%Y")

        lead_lines = []
        for i, lead in enumerate(new_leads, 1):
            line = f"{i}. <b>{lead['name']}</b> — {lead['type']}, {lead['zone']}"
            if lead.get("phone"):
                line += f"\n    📞 {lead['phone']}"
            if lead.get("website"):
                line += f"\n    🌐 {lead['website']}"
            if lead.get("rating"):
                line += f"\n    ⭐ {lead['rating']} ({lead.get('reviews', '?')} rec.)"
            lead_lines.append(line)

        header = f"🍷 <b>Retrogusto Lead Scout — {today_str}</b>\nTrovati <b>{len(new_leads)} nuovi lead</b>:\n"
        chunks = []
        current = header
        for line in lead_lines:
            if len(current) + len(line) + 2 > 4000:
                chunks.append(current)
                current = line + "\n"
            else:
                current += line + "\n"
        if current.strip():
            chunks.append(current)

        csv_path = save_leads_to_csv(new_leads)

        for idx, chunk in enumerate(chunks):
            if idx == len(chunks) - 1:
                chunk += "\n\n📎 CSV con i nuovi lead allegato."
            await send_telegram_message(client, chunk)
            await asyncio.sleep(0.5)

        await send_telegram_file(client, csv_path, f"Nuovi lead Retrogusto — {len(new_leads)} — {today_str}")

    log.info("✅ Completato")


if __name__ == "__main__":
    asyncio.run(main())
